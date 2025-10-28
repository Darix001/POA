import ujson
import unicodedata
import dataclasses
import operator as op
import itertools as it

from openpyxl.cell import Cell
from Levenshtein import distance
from openpyxl import load_workbook
from collections import defaultdict
from functools import partial, cache
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


Comment = partial(Comment, author="Dariel Buret")

verbs : dict[tuple, str] = {
	("Gasto", "Encuentro", "Inventario", "Eucaristia",
		 "Reunión", "Proyecto"):"Realizar",
	("Bienvenida","Seguimiento"):"Dar",
	("Olimpiada", "Día"):"Celebrar",
	"Levantamiento":"Levantar", "Uso":"Usar", "Gestión":"Gestionar",
}
ones = it.repeat(1)
'''
Cuando una acción despúes de la 1era palabra sea alguna de estas, 
el resto de palabras será igual a resto[valor:]

'''
rest_sl = {"de ":3, "del ":1}

infinitives = tuple(map("{}r".format, "aeiou"))

corregir_acciones = "Corregir Verbos de las acciones (donde sea posible)"
M = "MATRIZ"
L = "INSUMOS"
pres_col = "Presupuesto"
idcol = "Identificación"
act_col = "Acciones"
pa_col = "Presupuesto por acciones"
ff_col = "Fuente de Financiamiento"
prod_col = "Producto"
INPUT_FOLDER = "Pendientes"
OUTPUT_FOLDER = "Out"
utf_op = partial(open, encoding='utf-8')
misses:dict[str,int]
insumo:dict[str, dict[str,int]]

get_stem = op.attrgetter('stem')
cell_val = op.attrgetter("value")
dirs = op.attrgetter("bottom", "top", "right", "left")
get_color = op.attrgetter("color")
first_item = op.itemgetter(0)

#----------CELL BACKGROUND FILL COLORS ----------


fills = {"red":"FFFF0000", "yellow":"FFFFFF00"}

for color, code in fills.items():
	fills[color] = PatternFill(
		start_color=code,
		end_color=code,
		fill_type="solid"
		)

def normalize(string:str, /):
	'''returns a caseless unaccented version of the string'''
	# Normalize to NFD form (decomposes accents)
	string = unicodedata.normalize("NFD", string)
    # Remove diacritic marks
	selectors = map("Mn".removesuffix, map(unicodedata.category, string))
    # Case-insensitive comparison using casefold
	return "".join(map(str.casefold, it.compress(string, selectors)))


@dataclasses.dataclass(slots=True, frozen=True)
class Table:
	min_row:int
	max_row:int
	stmt_row:int
	headers_row:int
	id_num:int
	activities:list
	sheet:Worksheet

	def __len__(self, /):
		return self.max_row - self.min_row + 1

	def __str__(self, /):
		sheet = self.sheet
		start_col = get_column_letter(sheet.min_column)
		end_col = get_column_letter(sheet.max_column)
		return (
			f"{sheet.title}!{start_col}{self.min_row}:"
			f"{end_col}{self.max_row}"
			)

	def __getitem__(self, col, /):
		cols = self.sheet.act_cols
		if index := cols.get(col):
			col = index
		else:
			for col_name, index in cols.items():
				if col in col_name:
					cols[col] = col = index
					break
			else:
				raise KeyError("No such column", col)

		return self.column(col.column)

	def __getattr__(self, col):
		sheet = self.sheet
		if cell := sheet.stmt_cols.get(col):
			return sheet.cell(row=self.stmt_row, column=cell.column)
		else:
			raise AttributeError(col)


	def rows(self, /, min_col=None, max_col=None, values_only=False):
		sheet = self.sheet
		return sheet.iter_rows(
			min_row=self.min_row,
			max_row=self.max_row,
			min_col=min_col or 2,
			max_col=max_col,
			values_only=values_only)

	def column(self, /, col:str|int, values_only=False):
		return next(self.sheet.iter_cols(
			min_row=self.min_row,
			max_row=self.max_row,
			min_col=col,
			max_col=col,
			values_only=values_only), ())


def fx_sum(coord1, coord2, /, cl = None) -> str:
	if cl:
		coord1 = f"{cl}{coord1}"
		coord2 = f"{cl}{coord2}"
	return f"=SUM({coord1}:{coord2})"



def stw_key(string:str, stdict:dict) -> str|tuple|None:
	#Returns the first key that returns True for string.startswith(key)
	return next(filter(string.startswith, stdict), None)


def strip_nonids(string:str) -> int:
	#Strip all chars that are not identifiers
	start = next(it.compress(it.count(), 
			map(str.isidentifier, string)), len(string))
	return string[start:].rstrip(". ")


def col_names(row, start=2, key=first_item):
	return {v.strip().lower():cell for cell in row
		if type(v := cell.value) is str}


class Insumos_Group(dict):
	__slots__ = ()
	
	def __missing__(self, key:str, /):
		self[key] = value = [v for k,v in self.items() 
			if distance(key, k) <= 4 or key in k]
		return value
		


def scan_insumos(sheet, misses):
	alt_names = dict(zip(misses, map(Insumos_Group, misses.values())))
	col = iter(next(sheet.iter_cols(1,1, values_only=True)))
	data = defaultdict(Insumos_Group)
	start = op.indexOf(col, idcol) + 1

	for n, value in enumerate(col, start + 1):
		if value:
			value = normalize(value)
			data[value[0]][value] = n

	@cache
	def locate_insumo(insumo, /) -> str|list:
		insumo = normalize(insumo)
		if group := data.get(letter := insumo[0]):
			if result := group[insumo]:
				return f"=INSUMOS!A{result}" if type(result) is int else result
			else:
				if group := alt_names.get(letter):
					return group[insumo]

	return locate_insumo


def get_tables(sheet, pia={}, correct={}):
	#Delimitar las tablas de insumos y actividades
	tables = []
	acts = []
	stmt_row = prod_cols = cols = headers_row = None
	r = start = 0
	stop = 1
	enum_prod = all(map(correct.get, ("producto", "enumeraciones")))

	for id_num in it.count(1):
		stop = None
		rows = it.pairwise(sheet.iter_rows(r + 1, min_col=2))
		for row1, row2 in rows:
			first_cell = row1[0]
			if value := first_cell.value:
				value = value.strip().lower()
			else:
				continue

			if value.startswith("resultado") and ("intermedio" in value or \
				"efecto" in value):
				if not prod_cols:
					sheet.stmt_cols = prod_cols = col_names(row1)
					prod_cols["resultado"] = prod_cols.pop(
						next(iter(prod_cols)))


				if rng := sheet.get_merge_range(first_cell):
					stmt_row = rng.max_row + 1
				else:
					stmt_row = row2[0].row

				pcell = sheet.cell(row=stmt_row,
					column=prod_cols["producto"].column)

				prod_str = strip_nonids(pcell.value)
				
				if enum_prod:
					pcell.value = f"{id_num}. {prod_str}."

				acts = pia.get(prod_str, ())
				break

			elif value == "actividades y sus atributos":
				pass
		else:
			break

		for row1, row2 in rows:
			if (value := row1[0].value):
				value = value.strip().lower()
			else:
				continue
			
			if value == "actividades y sus atributos":
				headers_row = (r2cell := row2[0]).row
				start = r = sheet.get_merge_range(r2cell).max_row
				if not cols:
					sheet.act_cols = cols = col_names(row2)
				break
		else:
			break


		if start is None:
			break


		for cell in next(sheet.iter_cols(2, 2, r)):
			if type(cell) is Cell and \
				next(filter(get_color, dirs(cell.border)), None) is None:
				stop = (r := cell.row) - 1
				break

		tables.append(Table(
				min_row=start,
				max_row=stop or sheet.max_row,
				sheet=sheet,
				id_num=id_num,
				headers_row=headers_row,
				stmt_row=stmt_row,
				activities=acts))

	
		
	return tables


def get_merge_range(sheet, cell, /):
	coord = cell.coordinate
	for rng in sheet.merged_cells.ranges:
		if coord in rng:
			return rng

Worksheet.get_merge_range = get_merge_range


def pyxl_process(path, out, config):
	global all_insumos
	wb = load_workbook(path)
	sheetname = None
	for name in wb.sheetnames:
		if name.strip().upper().startswith("MATRIZ"):
			matriz = wb[name]
			sheetname = name
			break

	locate_insumo = scan_insumos(wb["INSUMOS"], config["misses"])
	tables = get_tables(matriz, config["Matriz PIA"], 
		correct := config["correct"])
	cols = matriz.act_cols
	prod_cols = matriz.stmt_cols

	ids = defaultdict(partial(it.count, 1))
	for table in tables:
		acts_col, acc_col = table["actividades"], table["acciones"]
		acts = [*table.activities]
		actv_id = None
		for actividad, accion in zip(acts_col, acc_col):
			if not (acc := accion.value) or acc.startswith("="):
				continue

			#----------ACTIVIDADES-------
			if correct["actividades"] and (actv := actividad.value):
				actv = strip_nonids(actv)
				actv_id = next(ids["actv"])
				ids.pop("acc", None)

				if actv not in acts:
					actividad.fill = fills["red"]
				else:
					acts.remove(actv)

				if correct["enumeraciones"]:
					actv = f"{actv_id}. {actv}."
				else:
					actv += '.'

				actividad.value = actv


			#----------ACCIONES-------
			if not correct["acciones"]:
				continue
			
			acc = strip_nonids(acc).capitalize() or "Insertar Acción"
			
			acc_parts = acc.split(maxsplit=1)
			first_word, rest = acc_parts

			if first_word.endswith(("ción", "cion")):
				first_word = first_word[:-4] + 'r'
				
				if key := stw_key(rest, rest_sl):
					rest = rest[rest_sl[key]:]

				acc_parts = [first_word, rest]

			elif not first_word.endswith(infinitives):
				if key := stw_key(first_word, verbs):
					acc_parts[0] = verbs[key]
				else:
					accion.fill = fills["red"]

			if correct["enumeraciones"]:
				acc_parts.insert(0,
					f"{table.id_num}.{actv_id}.{next(ids["acc"])}")

			accion.value = ' '.join(acc_parts)


		ids.clear()
		if acts:
			comment = Comment("Faltan Actividades:\n " + ",\n".join(acts))
			matriz.cell(
				row=table.headers_row,
				column=acts_col[0].column
				).comment = comment
	rng = matriz.get_merge_range(ins_cell := cols["insumos"])


	#Escanear Columna de identificación de insumos
	insumos = it.chain.from_iterable(
		map(op.methodcaller("column", ins_cell.column), tables)
		)

	for cell in insumos:
		if (value := cell.value) and value[0].isalpha() \
		 and (value := locate_insumo(value)):
			if type(value) is list:
				cell.comment = Comment(
					f"Posibles opciones: {',\n'.join(map(str, value))}")
			else:
				cell.value = value
				continue

			cell.fill = fills["red"]
		

	max_col = rng.max_col
	min_col = rng.min_col + 2
	ncols = (max_col - min_col + 1)
	months = (ncols - 2) // 2

	#Corregir fórmulas de cantidades y montos
	if correct["fórmulas"]:
		amount_cl = get_column_letter(max_col)
		
		for table in tables:
			col = filter(Cell.__instancecheck__,
			 table["presupuesto por acciones"])

			for fcell, lcell in it.pairwise(col):
				fcell.value = fx_sum(fcell.row, lcell.row - 1, amount_cl)

			ppa_cl = fcell.column_letter

		for table in tables:
			for row in table.rows(min_col, max_col):
				if not row[-1].value:
					continue

				batchs = it.batched(row := iter(row), months)
				quants = next(batchs)
				price = next(row)
				amounts = next(batchs)
				total_amount = next(row)
				
				price_coord = price.coordinate

				for quant, amount in zip(quants, amounts):
					amount.value = f"={quant.coordinate}*{price_coord}"
				
				#Sum of totals
				total_amount.value = fx_sum(
					amounts[0].coordinate, amounts[-1].coordinate)

			table.presupuesto.value = fx_sum(table.min_row,
				table.max_row, ppa_cl)

	
	if correct["resumen"]:
		prodcl = prod_cols["producto"].column_letter
		resumen = wb.create_sheet("RESUMEN PRESUPUESTAL")
		resumen.append(("Producto", "Presupuesto"))

		#Corregir Presupuesto por Producto
		if ' ' in sheetname:
			sheetname = f"'{sheetname}'"

		preffix = f"={sheetname}!"
		for table in tables:
			presupuesto = table.presupuesto
			resumen.append((
				f"{preffix}{prodcl}{presupuesto.row}",
				preffix + presupuesto.coordinate
				))

		resumen.append(("Total:", fx_sum(2, p := len(tables) + 1, "B")))
		resumen.append(("Presupuesto:", "0.0"))
		resumen.append(("Diferencia:", f"=B{p + 2}-B{p + 1}"))

	wb.save(out)
	wb.close()


def matriz_pia():
	wb = load_workbook(r"C:\Users\Daisy Garcia\Downloads\16 09 2025  MATRIZ COSTEO POA 2026 CENTRO EDUCATIVO.xlsm")
	sheet = wb["MATRIZ A TRABAJAR"]
	products = defaultdict(list)

	tables = get_tables(sheet)

	for table in tables:
		actividades = products[strip_nonids(table.producto.value)]
		actividades.extend(
			map(strip_nonids,
				filter(None,
					map(cell_val, table["actividades"])
					)
				)
			)

	print(ujson.dumps(products, indent=4, ensure_ascii=False))


if __name__ == '__main__':
	pass